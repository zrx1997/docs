# coding:utf-8
# 邮箱服务器需要支持SMTP和POP3协议
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText  # 邮件正文
from email.header import Header  # 邮件头部

# 加载excel文件
wb = load_workbook("员工工资条.xlsx", data_only=True)    # 员工工资条文件，可以写绝对路径

# 获取当前表格内容
sheet = wb.active
# print(sheet)

# 登录邮箱
smtp_obj = smtplib.SMTP_SSL("smtp.qq.com", 465)  # 邮箱服务器，端口
smtp_obj.login("xxxxx@qq.com", "xxxxxxxxxxxxx")  # 邮箱账号，授权码或者密码

# 生成工资条
count = 0
table_col_html = '<thead>'    # 表头
for row in sheet.iter_rows(min_row=1):
    count += 1
    if count == 1:
        for col in row:
            table_col_html += f"<th>{col.value}</th>"
        table_col_html += "</thead>"
        continue
    else:
        row_text = "<tr>"  # 开始一行
        for cell in row:
            row_text += f"<td>{cell.value}</td>"
        row_text += "</tr>"  # 结束一行
        name = row[4]
        staff_email = row[1].value

    # 设置邮件内容  &nbsp;代表空格
    mail_context = f'''
        <h3>{name.value},您好：</h3>
        <p>&nbsp;&nbsp;&nbsp;&nbsp;请查收你的工资条，若有疑问，请及时联系财务部，谢谢合作！
        <table border="1px solid black">
            {table_col_html}
            {row_text}
        </table>
        </p>
    '''
    msg_context = MIMEText(mail_context, "html", "utf-8")   # 指定邮箱正文的格式
    msg_context["From"] = Header("xx公司人事部<xxxxxxxx@qq.com>", "utf-8")  # 发送者相关信息
    msg_context["To"] = Header("xx公司员工", "utf-8")  # 接收者
    msg_context["Subject"] = Header("xx公司7月份员工工资条", "utf-8")  # 邮件主题

    # 发送邮件
    smtp_obj.sendmail("xxxxxxxxx@qq.com", [staff_email, ], msg_context.as_string())
    print(f"成功发送工资条到{staff_email}==>{name.value}.....")  # 打印发送结果
	
